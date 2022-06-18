from matplotlib.image import imread
import matplotlib.pyplot as plt
import numpy as np
import os
import cv2
from skimage.metrics import mean_squared_error
from skimage.metrics import structural_similarity
from skimage.util import random_noise
from sewar.full_ref import vifp
from openpyxl import Workbook, load_workbook

'''
	def read_image(args)

	funkcja służąca do wczytywania zdjęcia


	argumenty:
		path - ścieżka do pliku

	zwraca:
		zdjęcie jako dwuwymiarowa tablica (numpy array)
		lub zero, jeśli nie udało się wczytać zdjęcia
'''
def read_image(path=None):
	if path is None:
		print("Nie podano argumentu ze ścieżką zdjęcia")

		return None
	
	if not os.path.exists(path):
		print("Nie istnieje podany plik")

		return None
		
	return cv2.imread(path, cv2.IMREAD_GRAYSCALE)



'''
	def show_image(args)

	funkcja służąca do wyświetlania zdjęcia


	argumenty:
		label - etykieta zdjęcia
		img - zdjęcie do wyświetlenia
'''
def show_image(label="", img=None):
	if img is None:
		print("None")
	else:
		cv2.imshow(label, img)
		cv2.waitKey(0)
		cv2.destroyAllWindows()


'''
	def create_noise(args)

	funkcja nanosi szum na zdjęcie

	argumenty:
		image - zdjęcie, na które ma zostać naniesiony szum
		noise_mode - wybranie rodzaju szumu

	zwraca:
		zdjęcie z naniesionym szumem
'''
def create_noise(image, noise_mode):
	if noise_mode == 'gauss':
		return random_noise(image, mode='gaussian', var=0.01, mean=0.1)
	if noise_mode == 'poisson':
		return random_noise(image, mode='poisson')
	if noise_mode == 'pepper':
		return random_noise(image, mode="pepper", amount=0.2)
	if noise_mode == 's&p':
		return random_noise(image, mode="s&p", amount=0.2, salt_vs_pepper=0.2)
	if noise_mode == 'speckle':
		return random_noise(image, mode="speckle", var=0.02)


'''
	def metrics(args)

	funkcja oblicza wartości poszczególnych metryk


	argumenty:
		image - oryginalne zdjęcie
		image_approx - zaszumione aproksymowane zdjęcie

	zwraca:
		metryki mse, ssim i vif
'''
def metrics(image, image_approx):
	mse = mean_squared_error(image, image_approx)
	ssim = structural_similarity(image, image_approx, data_range = image_approx.max() - image_approx.min())
	vif = vifp(image, image_approx)

	return mse, ssim, vif

'''
	def wb_init(args)

	funkcja otwiera arkusz excel i nanosi w nim poprawiające czytelność nazwy kolumn

	argumenty:
		wb_name - ścieżka do arkusza

	zwraca:
		workbook i worksheet excel
'''

def wb_init(wb_name):
	wb = load_workbook(wb_name)
	ws = wb.active
	ws['A'+ str(2)].value = 'Rząd najlepszej macierzy aproksymacji względem maksymalnego rzędu zdjęcia.'
	ws['B'+ str(3)].value = 'MSE'
	ws['C'+ str(3)].value = 'SSIM'
	ws['D'+ str(3)].value = 'VIF'

	return wb, ws

'''
	def make_plots(args)

	funkcja tworzy wykresy poszczególnych metryk w zależności od rzędu aproksymacji

'''

def make_plots(MSE, best_rank_mse, SSIM, best_rank_ssim, VIF, best_rank_vif, ranks, noise):
	plt.rcParams['figure.figsize'] = [10, 6]
	plt.figure(1)
	plt.semilogx(ranks, MSE)
	plt.title('Błąd średniokwadratowy w zależności od rzędu macierzy aproksymującej (szum ' + str(noise) + ')')
	plt.xlabel('Rząd')
	plt.ylabel('MSE')
	r ="r = %.2f %%" % best_rank_mse
	m =' MSE = %.2f' % min(MSE)
	plt.annotate(r + m, (ranks[MSE.index(min(MSE))], min(MSE)))
	plt.show()

	plt.figure(1)
	plt.semilogx(ranks, SSIM)
	plt.title('SSIM w zależności od rzędu macierzy aproksymującej (szum ' + str(noise) + ')')
	plt.xlabel('Rząd')
	plt.ylabel('SSIM')
	r ="r = %.2f %%" % best_rank_ssim
	m =' SSIM = %.4f' % max(SSIM)
	plt.annotate(r + m, (ranks[SSIM.index(max(SSIM))], max(SSIM)))
	plt.show()

	plt.figure(1)
	plt.semilogx(ranks, VIF)
	plt.title('VIF w zależności od rzędu macierzy aproksymującej (szum ' + str(noise) + ')')
	plt.xlabel('Rząd')
	plt.ylabel('VIF')
	r ="r = %.2f %%" % best_rank_vif
	m =' VIF = %.4f' % max(VIF)
	plt.annotate(r + m, (ranks[VIF.index(max(VIF))], max(VIF)))
	plt.show()


if __name__ == "__main__":
	wb_name = 'opsi.xlsx'		# ścieżka do arkusza excel
	wb, ws = wb_init(wb_name)
	cell_row = 4

	# ścieżki zdjęć i nazwy szumów, które mają zostać poddane analizie
	# images = ['ct.jpg', 'ct1.png', 'ct2.png', 'mr.jpg', 'mr1.png', 'mr2.png', 'rtg1.png', 'rtg2.png', 'rtg3.png', 'usg1.png', 'usg2.png', 'usg3.png', 'figury.png']
	# noises = ['gauss', 'poisson', 'pepper', 's&p', 'speckle']
	images = ['ct.jpg']
	noises = ['gauss']

	for image in images:
		ws['A'+ str(cell_row)].value = image
		img = read_image(image)
		for noise in noises:
			img_noise = create_noise(img, noise)
			img_noise *= 255

			U, S, VT = np.linalg.svd(img_noise, full_matrices=False)
			S = np.diag(S)
			rank = np.linalg.matrix_rank(S)								# odczytanie maksymalnego rzędu zdjęcia
			ranks = np.around(np.geomspace(3, rank, num=50, endpoint=True)).astype(int)	# tworzenie wektora rzędów, dla których wyznaczane będą metryki, w sposób logarytmiczny 

			MSE = []
			SSIM = []
			VIF = []
			for r in ranks:
				img_approx = U[:, :r] @ S[:r, :r] @ VT[:r, :]
				mse, ssim, vif = metrics(img, img_approx)
				MSE.append(mse)
				SSIM.append(ssim)
				VIF.append(vif)
			best_rank_mse = ranks[MSE.index(min(MSE))]*100/rank	
			best_rank_ssim = ranks[SSIM.index(max(SSIM))]*100/rank
			best_rank_vif = ranks[VIF.index(max(VIF))]*100/rank

			# zapisanie wyników do arkusza
			ws['B'+ str(cell_row)].value = best_rank_mse/100
			ws['C'+ str(cell_row)].value = best_rank_ssim/100
			ws['D'+ str(cell_row)].value = best_rank_vif/100
			cell_row += 1

			make_plots(MSE, best_rank_mse, SSIM, best_rank_ssim, VIF, best_rank_vif, ranks, noise)
		cell_row += 2

	wb.save(wb_name)